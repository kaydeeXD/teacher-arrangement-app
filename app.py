import streamlit as st
import pandas as pd
import random
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# === UI STATE CONTROL ===
if "show_suggestions_panel" not in st.session_state:
    st.session_state.show_suggestions_panel = False

def toggle_suggestions_panel():
    st.session_state.show_suggestions_panel = not st.session_state.show_suggestions_panel

# === SESSION STATE FOR FILE ===
if "uploaded_file" not in st.session_state:
    st.session_state.uploaded_file = None

if "weekly_arrangements" not in st.session_state:
    st.session_state.weekly_arrangements = []

# === HEADER WITH LOGO, SCHOOL NAME, DATE ===
today = datetime.today().strftime("%A, %d %B %Y")
st.image("KV logo.png", use_container_width =True)
st.markdown(f"""
    <div style='text-align: center; padding: 10px;'>
        <h1 style='color: #1f4e79;'>🧑‍🏫 Teacher Arrangement System</h1>
        <h2 style='color: #003366; margin-top: -10px;'>Kendriya Vidyalaya Kishtwar</h2>
        <h4 style='margin-top: -5px; color: gray;'>Date: {today}</h4>
        <hr style='margin-top: 15px; margin-bottom: 25px;'>
    </div>
""", unsafe_allow_html=True)

# === SIDEBAR ===
st.sidebar.title("Teacher Arrangement Generator")
file_input = st.sidebar.file_uploader("Upload updated 'KV TT.xlsx'", type=["xlsx"])
selected_day = st.sidebar.selectbox("Select Day", ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"])

if file_input:
    st.session_state.uploaded_file = file_input
    st.sidebar.success("✅ Uploaded file in use.")
elif st.session_state.uploaded_file is not None:
    file_input = st.session_state.uploaded_file
    st.sidebar.info("ℹ️ Using previously uploaded file.")
else:
    try:
        file_input = "KV TT.xlsx"
        st.sidebar.info("ℹ️ Using default file: 'KV TT.xlsx'")
    except FileNotFoundError:
        st.sidebar.error("❌ No file uploaded and default file not found.")
        st.stop()

# === PARSE TIMETABLE ===
@st.cache_data
def parse_timetable(file):
    df = pd.read_excel(file, sheet_name="TEACHER  WISE", header=None)
    parsed_rows = []
    current_teacher = None

    for _, row in df.iterrows():
        first_cell = str(row[0]).strip() if pd.notna(row[0]) else ""
        if first_cell and first_cell.upper() not in ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"] \
                and not first_cell.isdigit() and "TOTAL" not in first_cell.upper() and "TPOD" not in first_cell.upper():
            current_teacher = first_cell
            continue
        if first_cell.upper() in ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"] and current_teacher:
            day = first_cell.capitalize()
            tpod_val = row[9] if pd.notna(row[9]) else None
            for period_num in range(1, 9):
                class_val = row[period_num] if pd.notna(row[period_num]) else None
                parsed_rows.append({
                    "Teacher": current_teacher,
                    "Day": day,
                    "Period": period_num,
                    "Class": str(class_val).strip() if class_val else None,
                    "TPOD": int(tpod_val) if pd.notna(tpod_val) else None
                })

    df = pd.DataFrame(parsed_rows)
    misc_keywords = ["PH&E", "YOGA TEACHER", "SPORTS COACH", "ART", "DRAWING", "COMPUTER INSTRUCTOR", "LIBR.", "WET", "MUSIC"]

    def get_teacher_domain(name):
        name_upper = name.upper()
        if any(k in name_upper for k in misc_keywords): return "Misc"
        elif "PRINCIPAL" in name_upper: return "Principal"
        elif "PGT" in name_upper: return "PGT"
        elif "TGT" in name_upper: return "TGT"
        elif "PRT" in name_upper: return "PRT"
        else: return "Unknown"

    df["Domain"] = df["Teacher"].apply(get_teacher_domain)
    return df

def extract_class_level(class_str):
    roman_map = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6, "VII": 7, "VIII": 8, "IX": 9, "X": 10, "XI": 11, "XII": 12}
    if not class_str or not isinstance(class_str, str): return None
    clean_str = str(class_str).upper().strip()
    first_part = clean_str.split()[0]
    return roman_map.get(first_part, None)

# === DAY MODE SELECTION ===
day_mode = st.radio(
    "Select Day Mode",
    options=["Full Day", "Half Day", "Custom Periods"],
    horizontal=True
)

if day_mode == "Full Day":
    selected_periods = list(range(1, 9))
elif day_mode == "Half Day":
    selected_periods = list(range(1, 5))
else:
    all_periods = [f"Period {i}" for i in range(1, 9)]
    custom = st.multiselect("Select specific periods for arrangement", options=all_periods, default=[])
    selected_periods = [int(p.split()[1]) for p in custom]

def generate_arrangement(absent_dict, day, timetable_df):
    arrangements = []
    suggested_arrangements = []
    arrangement_count = {}
    arrangement_tracker = {}
    day_df = timetable_df[timetable_df["Day"].str.lower() == day.lower()]

    for absent_teacher, absence_type in absent_dict.items():
        if absence_type == "1st half":
            teacher_schedule = day_df[(day_df["Teacher"] == absent_teacher) & (day_df["Period"].isin(range(1, 5)))]
        elif absence_type == "2nd half":
            teacher_schedule = day_df[(day_df["Teacher"] == absent_teacher) & (day_df["Period"].isin(range(5, 9)))]
        else:
            teacher_schedule = day_df[(day_df["Teacher"] == absent_teacher)]

        teacher_schedule = teacher_schedule[teacher_schedule["Period"].isin(selected_periods)]

        for _, row in teacher_schedule.iterrows():
            target_class = row["Class"]
            if pd.isna(target_class) or str(target_class).strip().upper() in ["", "CCA", "LIB", "LIBRARY", "P.E.", "SPORTS"]:
                continue

            period = row["Period"]
            level = extract_class_level(target_class)
            if level is None:
                continue

            target_domain = "Primary" if level <= 5 else "Secondary" if level <= 10 else "Senior Secondary"

            free_teachers = day_df[
                (day_df["Period"] == period) &
                (~day_df["Teacher"].isin(absent_dict.keys())) &
                (day_df["TPOD"] < 7) &
                (day_df["Class"].isna() | day_df["Class"].isin(["", "CCA", "LIB", "LIBRARY", "P.E.", "SPORTS"]))
            ].copy()

            substitute = None
            suggested_teachers = []
            domain_priority = {
                "Senior Secondary": ["PGT", "Principal", "Misc"],
                "Secondary": ["TGT", "Misc", "PGT", "Principal"],
                "Primary": ["PRT", "Misc", "Principal"]
            }

            candidates = pd.DataFrame()
            for domain in domain_priority[target_domain]:
                candidates = free_teachers[free_teachers["Domain"] == domain]
                if not candidates.empty:
                    break

            if candidates.empty:
                relaxed_free = day_df[
                    (day_df["Period"] == period) &
                    (~day_df["Teacher"].isin(absent_dict.keys())) &
                    (day_df["Class"].isna() | day_df["Class"].isin(["", "CCA", "LIB", "LIBRARY", "P.E.", "SPORTS"]))
                ]
                for domain in domain_priority[target_domain]:
                    candidates = relaxed_free[relaxed_free["Domain"] == domain]
                    if not candidates.empty:
                        break

            if not candidates.empty:
                teacher_list = list(candidates["Teacher"].unique())
                random.shuffle(teacher_list)
                teacher_list.sort(key=lambda t: arrangement_count.get(t, 0))

                suggested_teachers = teacher_list.copy()

                for t in teacher_list:
                    if arrangement_tracker.get((t, period), False):
                        continue
                    substitute = t
                    arrangement_tracker[(t, period)] = True
                    arrangement_count[substitute] = arrangement_count.get(substitute, 0) + 1
                    break

            arrangements.append({
                "Absent Teacher": absent_teacher,
                "Period": period,
                "Class": target_class,
                "Substitute Teacher": substitute
            })

            suggested_arrangements.append({
                "Absent Teacher": absent_teacher,
                "Period": period,
                "Class": target_class,
                "Suggested Teachers": ", ".join(suggested_teachers[:5]) if suggested_teachers else ""
            })

    df = pd.DataFrame(arrangements)
    df["Sub_with_Class"] = df.apply(
        lambda x: f"{x['Substitute Teacher']} ({x['Class']})" if pd.notna(x['Substitute Teacher']) else "", axis=1
    )
    pivot_df = df.pivot(index="Absent Teacher", columns="Period", values="Sub_with_Class").fillna("")
    for p in selected_periods:
        if p not in pivot_df.columns:
            pivot_df[p] = ""
    pivot_df = pivot_df[[p for p in selected_periods]]
    pivot_df.columns = [f"Period {p}" for p in pivot_df.columns]
    output_df_reset = pivot_df.reset_index()
    output_df_reset.insert(1, "Reason", output_df_reset["Absent Teacher"].map(absence_reason_dict))

    suggestions_df = pd.DataFrame(suggested_arrangements)

    st.session_state.weekly_arrangements.append({
        "date": today,
        "day": day,
        "arrangement": output_df_reset
    })

    st.session_state["generated_arrangement"] = output_df_reset
    st.session_state["suggestions_df"] = suggestions_df

    return output_df_reset, suggestions_df


# === MAIN LOGIC ===
timetable_df = parse_timetable(file_input)
teacher_list = timetable_df["Teacher"].unique().tolist()
absent_teachers = st.multiselect("Select Absent Teachers", teacher_list)
absence_types = ["Full", "1st half", "2nd half"]

absent_dict = {}
absence_reason_dict = {}

for t in absent_teachers:
    col1, col2 = st.columns([1, 2])
    with col1:
        absent_dict[t] = st.selectbox(f"{t} absence type", options=absence_types, key=f"{t}_absence_type")
    with col2:
        absence_reason_dict[t] = st.text_input(f"{t} absence reason", placeholder="Enter reason", key=f"{t}_reason_input")
# === Show Final Arrangement Table if It Exists in Session State ===
if "generated_arrangement" in st.session_state:
    st.subheader("📋 Arrangements")
    st.dataframe(st.session_state["generated_arrangement"], use_container_width=True)
if st.button("🚀 Generate Arrangement"):
    output_df, suggestions_df = generate_arrangement(absent_dict, selected_day, timetable_df)
    st.success("✅ Arrangement Generated")
    st.subheader("📋 Arrangements")
    st.dataframe(output_df, use_container_width=True)
    # st.markdown("### 💡 Suggested Teachers")
    # st.dataframe(suggestions_df, use_container_width=True)

    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=output_df.shape[1])
    ws.cell(row=1, column=1).value = f"Arrangement for {today}"
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    for c_idx, col_name in enumerate(output_df.columns, start=1):
        ws.cell(row=2, column=c_idx).value = col_name
        ws.cell(row=2, column=c_idx).font = Font(bold=True)

    for r_idx, row in enumerate(output_df.itertuples(index=False), start=3):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value

    wb.save(output)
    st.download_button("📥 Download Excel", data=output.getvalue(), file_name="arrangement.xlsx")

    # === CONFLICT CHECKER ===
    st.markdown("### ⚠️ Conflict Report")
    assigned_periods = {}
    conflict_rows = []

    # Explicitly get period columns only
    period_cols = [col for col in output_df.columns if col.startswith("Period")]

    for _, row in output_df.iterrows():
        teacher = row["Absent Teacher"]
        for idx, col in enumerate(period_cols, start=1):
            assigned = row[col]
            if pd.notna(assigned) and assigned.strip():
                sub_teacher = assigned.split(" (")[0].strip()
                key = (sub_teacher, idx)
                if key in assigned_periods:
                    conflict_rows.append({
                        "Conflict Period": f"Period {idx}",
                        "Teacher": sub_teacher,
                        "Conflicting With": assigned_periods[key],
                        "Also Assigned To": teacher
                    })
                else:
                    assigned_periods[key] = teacher

    if conflict_rows:
        conflict_df = pd.DataFrame(conflict_rows)
        st.error("🚨 Time-slot conflicts detected!")
        st.dataframe(conflict_df, use_container_width=True)
    else:
        st.success("✅ No time-slot conflicts detected.")

    # === ARRANGEMENT LOAD VISUALIZATION ===
    st.markdown("### 📊 Arrangement Load per Substitute Teacher")
    load_counter = {}

    period_cols = [col for col in output_df.columns if col.startswith("Period")]
    for _, row in output_df.iterrows():
        for col in period_cols:
            assigned = row[col]
            if pd.notna(assigned) and assigned.strip():
                sub_teacher = assigned.split(" (")[0].strip()
                load_counter[sub_teacher] = load_counter.get(sub_teacher, 0) + 1

    if load_counter:
        load_df = pd.DataFrame({
            "Teacher": list(load_counter.keys()),
            "Assigned Periods": list(load_counter.values())
        }).sort_values(by="Assigned Periods", ascending=False)

        st.bar_chart(load_df.set_index("Teacher"))
    else:
        st.info("No assignments to visualize.")

st.markdown("---")
st.markdown("### 🔍 Get Suggested Substitute Teachers")
st.button("🔄 Suggestion Panel", on_click=toggle_suggestions_panel)
if st.session_state.show_suggestions_panel:
    if "suggestions_df" in st.session_state:
        absent_teachers_list = st.session_state["suggestions_df"]["Absent Teacher"].unique().tolist()

        selected_absent = st.selectbox(
            "Select an absent teacher",
            absent_teachers_list,
            key="selected_suggestion_teacher"
        )

        filtered_suggestions = st.session_state["suggestions_df"][
            st.session_state["suggestions_df"]["Absent Teacher"] == selected_absent
        ]

        st.dataframe(filtered_suggestions, use_container_width=True)
    else:
        st.warning("⚠️ Please generate the arrangement first.")
        
# === MANUAL EDITING INTERFACE TRIGGER ===
if "generated_arrangement" in st.session_state and "suggestions_df" in st.session_state:
    st.markdown("---")
    st.markdown("### 🧑‍🏫 Manual Edit Timetable")

    if "edit_queue" not in st.session_state:
        st.session_state.edit_queue = []

    if st.button("➕ Add Teacher for Editing"):
        st.session_state.edit_queue.append({
            "teacher": "",
            "periods": [],
            "edits": {}
        })

    original_df = st.session_state["generated_arrangement"]
    suggestions_df = st.session_state["suggestions_df"]
    editable_df = original_df.copy()
    period_columns = [col for col in editable_df.columns if col.startswith("Period")]

    for idx, entry in enumerate(st.session_state.edit_queue):
        st.markdown(f"---\n#### 📝 Edit Entry #{idx + 1}")
        col1, col2, col3 = st.columns([4, 4, 1])

        with col3:
            if st.button(f"🗑️", key=f"delete_entry_{idx}"):
                st.session_state.edit_queue.pop(idx)
                st.rerun()

        with col1:
            absent_list = editable_df["Absent Teacher"].unique().tolist()
            selected_teacher = st.selectbox(
                f"👤 Absent Teacher (Entry #{idx + 1})",
                [""] + absent_list,
                index=absent_list.index(entry["teacher"]) + 1 if entry["teacher"] in absent_list else 0,
                key=f"teacher_{idx}"
            )
            entry["teacher"] = selected_teacher

        if selected_teacher:
            with col2:
                teacher_df = editable_df[editable_df["Absent Teacher"] == selected_teacher]
                teacher_periods = [int(col.split(" ")[1]) for col in period_columns if not teacher_df[col].isnull().all()]
                selected_periods = st.multiselect(
                    f"🕘 Periods for {selected_teacher}",
                    options=sorted(teacher_periods),
                    default=entry["periods"],
                    key=f"periods_{idx}"
                )
                entry["periods"] = selected_periods

            for period_num in selected_periods:
                period_col = f"Period {period_num}"
                current_value = teacher_df.iloc[0][period_col]
                current_teacher = current_value.split(" (")[0].strip() if current_value else ""

                suggestion = suggestions_df[
                    (suggestions_df["Absent Teacher"] == selected_teacher) &
                    (suggestions_df["Period"] == period_num)
                ]
                class_val = suggestion["Class"].values[0] if not suggestion.empty else "N/A"
                suggested_teachers = suggestion["Suggested Teachers"].values[0].split(", ") if not suggestion.empty else []
                options = [""] + suggested_teachers

                substitute = st.selectbox(
                    f"➡️ Substitute for Period {period_num} (Class: {class_val})",
                    options=options,
                    index=options.index(current_teacher) if current_teacher in options else 0,
                    key=f"{selected_teacher}_{period_num}_{idx}"
                )
                entry["edits"][period_num] = f"{substitute} ({class_val})" if substitute else ""

    if st.session_state.edit_queue:
        st.markdown("---")
        if st.button("✅ Save All Changes"):
            for entry in st.session_state.edit_queue:
                teacher = entry["teacher"]
                if not teacher: continue
                for period in entry["periods"]:
                    period_col = f"Period {period}"
                    new_val = entry["edits"].get(period, "")
                    editable_df.loc[editable_df["Absent Teacher"] == teacher, period_col] = new_val

            st.session_state["final_arrangement"] = editable_df
            st.success("✅ All changes saved successfully.")

            # Update arrangement to reflect changes
            st.session_state["generated_arrangement"] = editable_df
            # === Show Updated Timetable in Pivot Format ===
            st.markdown("### 🗂️ Updated Arrangement Timetable")
            pivot_df = editable_df.drop(columns=["Absent Teacher"])
            pivot_df.insert(0, "Class", editable_df["Absent Teacher"])  # Rename for display
            pivot_df.rename(columns={"Class": "Absent Teacher"}, inplace=True)
            st.dataframe(pivot_df, use_container_width=True)

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                editable_df.to_excel(writer, index=False, sheet_name="Updated Arrangement")

            st.download_button(
                label="📥 Download Updated Timetable",
                data=output.getvalue(),
                file_name="Updated_Arrangement.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # === Conflict and Load Check After Saving ===
            st.markdown("### ⚠️ Conflict Report")
            assigned_periods = {}
            conflict_rows = []
            period_cols = [col for col in editable_df.columns if col.startswith("Period")]

            for _, row in editable_df.iterrows():
                teacher = row["Absent Teacher"]
                for idx, col in enumerate(period_cols, start=1):
                    assigned = row[col]
                    if pd.notna(assigned) and assigned.strip():
                        sub_teacher = assigned.split(" (")[0].strip()
                        key = (sub_teacher, idx)
                        if key in assigned_periods:
                            conflict_rows.append({
                                "Conflict Period": f"Period {idx}",
                                "Teacher": sub_teacher,
                                "Conflicting With": assigned_periods[key],
                                "Also Assigned To": teacher
                            })
                        else:
                            assigned_periods[key] = teacher

            if conflict_rows:
                conflict_df = pd.DataFrame(conflict_rows)
                st.error("🚨 Time-slot conflicts detected!")
                st.dataframe(conflict_df, use_container_width=True)
            else:
                st.success("✅ No time-slot conflicts detected.")

            st.markdown("### 📊 Arrangement Load per Substitute Teacher")
            load_counter = {}
            for _, row in editable_df.iterrows():
                for col in period_cols:
                    assigned = row[col]
                    if pd.notna(assigned) and assigned.strip():
                        sub_teacher = assigned.split(" (")[0].strip()
                        load_counter[sub_teacher] = load_counter.get(sub_teacher, 0) + 1

            if load_counter:
                load_df = pd.DataFrame({
                    "Teacher": list(load_counter.keys()),
                    "Assigned Periods": list(load_counter.values())
                }).sort_values(by="Assigned Periods", ascending=False)

                st.bar_chart(load_df.set_index("Teacher"))
            else:
                st.info("No assignments to visualize.")

# === WEEKLY TRACKER VIEW ===
st.markdown("---")
st.markdown("### 📅 Weekly Arrangement Tracker")
if st.button("🧾 Show Weekly Arrangement"):
    if not st.session_state.weekly_arrangements:
        st.info("No arrangements generated this week.")
    else:
        for record in st.session_state.weekly_arrangements:
            st.markdown(f"### 📌 {record['day']} - {record['date']}")
            st.dataframe(record["arrangement"], use_container_width=True)
            st.markdown("---")
