from pathlib import Path
import pandas as pd
from datetime import datetime
from io import BytesIO
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from parser import parse_timetable
from arranger import generate_arrangement
from gsheet import get_or_create_worksheet, load_df_from_gsheet
from persistence import persist_weekly_log, save_state_to_sheet, load_state_from_sheet, load_weekly_log, append_to_monthly_log
from constants import SPREADSHEET_ID
from utils import is_same_week, get_current_week_dates, get_last_week_dates

# Initialize Streamlit app
st.set_page_config(page_title="Teacher Arrangement System", layout="wide")

# Load state from Google Sheets
PersistentStateWorksheet = get_or_create_worksheet(SPREADSHEET_ID, "PersistentState")
date_str, day_mode, absent_teachers, reasons_dict, custom_periods, final_timetable_df, suggestions_df = load_state_from_sheet(PersistentStateWorksheet)

# Session state initialization
if "show_suggestions_panel" not in st.session_state:
    st.session_state.show_suggestions_panel = False

if "uploaded_file" not in st.session_state:
    st.session_state.uploaded_file = None

if "weekly_arrangements" not in st.session_state:
    weekly_log_df = load_weekly_log(SPREADSHEET_ID)

    if not weekly_log_df.empty:
        grouped = weekly_log_df.groupby(["Date", "Day"])
        st.session_state.weekly_arrangements = []
        for (date, day), group_df in grouped:
            group_df = group_df.drop(columns=["Date", "Day"])
            st.session_state.weekly_arrangements.append({
                "date": date,
                "day": day,
                "arrangement": group_df.reset_index(drop=True)
            })
    else:
        st.session_state.weekly_arrangements = []

if "generated_arrangement" not in st.session_state:
    result = load_state_from_sheet(PersistentStateWorksheet)
    if result:
        date_str, day_mode, absent_teachers, reasons_dict, custom_periods, timetable_df, suggestions_df = result
        today_str = datetime.today().strftime("%A, %d %B %Y")
        if date_str == today_str:
            st.session_state["generated_arrangement"] = timetable_df
            st.session_state["suggestions_df"] = suggestions_df.copy()
            st.session_state["__meta__date"] = date_str
            st.session_state["__meta__day_mode"] = day_mode
            st.session_state["__meta__absent_teachers"] = absent_teachers
            st.session_state["__meta__reasons"] = reasons_dict
            st.session_state["__meta__custom_periods"] = custom_periods
            st.toast("✅ Previous session data restored.")
        else:
            PersistentStateWorksheet.clear()
            st.toast("⚠️ Outdated state found. Cleared stale data.")
    else:
        st.toast("⚠️ No session data found.")

if "__meta__custom_periods" not in st.session_state:
    st.session_state["__meta__custom_periods"] = []

# UI: Header
today = datetime.today().strftime("%A, %d %B %Y")
day = datetime.today().strftime("%A")
is_sunday = (day == "Sunday")
selected_day = day
logo_path = Path(__file__).parent.parent / "assets" / "KV logo.png"
st.image(str(logo_path), use_container_width=True)
st.markdown(f"""
    <div style='text-align: center; padding: 10px;'>
        <h1 style='color: #1f4e79;'>🧑‍🏫 Teacher Arrangement System</h1>
        <h2 style='color: #003366; margin-top: -10px;'>Kendriya Vidyalaya Kishtwar</h2>
        <h4 style='margin-top: -5px; color: gray;'>Date: {today}</h4>
        <hr style='margin-top: 15px; margin-bottom: 25px;'>
    </div>
""", unsafe_allow_html=True)

# Sidebar Navigation
st.sidebar.title("Navigation")
page = st.sidebar.radio("Go to", ["🏠 Home", "📊 Arrangement Tracker"])

# File upload (only shown on Home page)
if page == "🏠 Home":
    st.sidebar.title("Teacher Arrangement Generator")
    file_input = st.sidebar.file_uploader("Upload Timetable", type=["xlsx"])

    if file_input:
        st.session_state.uploaded_file = file_input
        st.sidebar.success("✅ Uploaded file in use.")
    elif st.session_state.uploaded_file is not None:
        file_input = st.session_state.uploaded_file
        st.sidebar.info("ℹ️ Using previously uploaded file.")
    else:
        file_path = Path(__file__).parent.parent / "assets" / "KV TT.xlsx"
        if file_path.exists():
            file_input = str(file_path)
            st.sidebar.info("ℹ️ Using default file: 'KV TT.xlsx'")
        else:
            st.sidebar.error("❌ No file uploaded and default file not found.")
            st.stop()

if page == "🏠 Home":
    if is_sunday:
        st.warning("⚠️ No new arrangements can be generated on Sundays. You may only view the last saved timetable.")
    else:
        # Day mode selection
        prev_day_mode = st.session_state.get("__meta__day_mode", "Full Day")
        day_mode = st.radio(
            "Select Day Mode",
            options=["Full Day", "Half Day", "Custom Periods"],
            horizontal=True,
            index=["Full Day", "Half Day", "Custom Periods"].index(prev_day_mode)
        )

        # Determine selected periods based on mode
        if day_mode == "Full Day":
            selected_periods = list(range(1, 9))
        elif day_mode == "Half Day":
            selected_periods = list(range(1, 5))
        else:
            all_periods = [f"Period {i}" for i in range(1, 9)]
            custom = st.multiselect(
                "Select specific periods for arrangement",
                options=all_periods,
                default=st.session_state.get("__meta__custom_periods", []),
                key="__meta__custom_periods"
            )
            selected_periods = [int(p.split()[1]) for p in custom]

        # Save selected day mode
        st.session_state["__meta__day_mode"] = day_mode

        # Load timetable data
        timetable_df = parse_timetable(file_input)
        teacher_list = timetable_df["Teacher"].unique().tolist()

        # Absence inputs
        prev_absent_teachers = st.session_state.get("__meta__absent_teachers", [])
        absent_teachers = st.multiselect("Select Absent Teachers", teacher_list, default=prev_absent_teachers)
        absence_types = ["Full", "1st half", "2nd half"]
        absent_dict = {}
        absence_reason_dict = {}

        # Load previously selected types and reasons if available
        prev_absent_teachers = st.session_state.get("__meta__absent_teachers", [])
        prev_reasons = st.session_state.get("__meta__reasons", {})

        for t in absent_teachers:
            col1, col2 = st.columns([1, 2])
            default_absence_type = "Full"
            if t in prev_absent_teachers:
                saved_type = prev_reasons.get(t, "")
                if saved_type in ["1st half", "2nd half", "Full"]:
                    default_absence_type = saved_type
            with col1:
                absent_dict[t] = st.selectbox(
                    f"{t} absence type",
                    options=absence_types,
                    index=absence_types.index(default_absence_type) if default_absence_type in absence_types else 0,
                    key=f"{t}_absence_type"
                )

            default_reason = prev_reasons.get(t, "")
            with col2:
                absence_reason_dict[t] = st.text_input(
                    f"{t} absence reason",
                    placeholder="Enter reason",
                    value=default_reason,
                    key=f"{t}_reason_input"
                )

        # Show Final Arrangement Table if It Exists in Session State
        if "generated_arrangement" in st.session_state:
            st.subheader(f"📋 {today} Arrangements")
            st.dataframe(st.session_state["generated_arrangement"], use_container_width=True)

        if st.button("🚀 Generate Arrangement"):
            output_df, suggestions_df = generate_arrangement(
                absent_dict, absence_reason_dict, selected_periods, selected_day,
                day_mode, PersistentStateWorksheet, timetable_df
            )
            st.success("✅ Arrangement Generated")
            st.subheader("📋 Arrangements")
            st.dataframe(output_df, use_container_width=True)

            # Update session state
            st.session_state["generated_arrangement"] = output_df
            st.session_state["suggestions_df"] = suggestions_df

            # Update weekly arrangements in session
            today_str = datetime.today().strftime("%A, %d %B %Y")
            day_str = datetime.today().strftime("%A")
            updated = False
            for i, entry in enumerate(st.session_state.weekly_arrangements):
                if entry["date"] == today_str:
                    st.session_state.weekly_arrangements[i] = {
                        "date": today_str,
                        "day": day_str,
                        "arrangement": output_df
                    }
                    updated = True
                    break

            if not updated:
                st.session_state.weekly_arrangements.append({
                    "date": today_str,
                    "day": day_str,
                    "arrangement": output_df
                })

            # Persist WeeklyLog
            weekly_log_df = pd.concat([
                log["arrangement"].assign(Date=log["date"], Day=log["day"])
                for log in st.session_state.weekly_arrangements
                if is_same_week(log["date"])
            ])
            try:
                persist_weekly_log(weekly_log_df, SPREADSHEET_ID)
                append_to_monthly_log(output_df, SPREADSHEET_ID)
                st.success("✅ Weekly and Monthly arrangement updated.")
            except Exception as e:
                st.error(f"❌ Failed to update WeeklyLog or MonthLog: {e}")

            # Download Excel
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=output_df.shape[1])
            ws.cell(row=1, column=1).value = f"Arrangement for {today_str}"
            ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)

            for c_idx, col_name in enumerate(output_df.columns, start=1):
                ws.cell(row=2, column=c_idx).value = col_name
                ws.cell(row=2, column=c_idx).font = Font(bold=True)

            for r_idx, row in enumerate(output_df.itertuples(index=False), start=3):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx).value = value

            wb.save(output)
            st.download_button(
                label="📥 Download Excel",
                data=output.getvalue(),
                file_name=f"arrangement_{today_str.replace(',', '').replace(' ', '_')}.xlsx"
            )
            
            # === CONFLICT CHECKER ===
            st.markdown("### ⚠️ Conflict Report")
            assigned_periods = {}
            conflict_rows = []

            # Explicitly get period columns only
            period_cols = [col for col in output_df.columns if col.startswith("Period")]

            for _, row in output_df.iterrows():
                teacher = row["Absent Teacher"]
                for idx, col in enumerate(period_cols, start=1):
                    assigned = row[col]
                    if pd.notna(assigned) and assigned.strip():
                        sub_teacher = assigned.split(" (")[0].strip()
                        key = (sub_teacher, idx)
                        if key in assigned_periods:
                            conflict_rows.append({
                                "Conflict Period": f"Period {idx}",
                                "Teacher": sub_teacher,
                                "Conflicting With": assigned_periods[key],
                                "Also Assigned To": teacher
                            })
                        else:
                            assigned_periods[key] = teacher

            if conflict_rows:
                conflict_df = pd.DataFrame(conflict_rows)
                st.error("🚨 Time-slot conflicts detected!")
                st.dataframe(conflict_df, use_container_width=True)
            else:
                st.success("✅ No time-slot conflicts detected.")

            # === ARRANGEMENT LOAD VISUALIZATION ===
            st.markdown("### 📊 Arrangement Load per Substitute Teacher")
            load_counter = {}

            period_cols = [col for col in output_df.columns if col.startswith("Period")]
            for _, row in output_df.iterrows():
                for col in period_cols:
                    assigned = row[col]
                    if pd.notna(assigned) and assigned.strip():
                        sub_teacher = assigned.split(" (")[0].strip()
                        load_counter[sub_teacher] = load_counter.get(sub_teacher, 0) + 1

            if load_counter:
                load_df = pd.DataFrame({
                    "Teacher": list(load_counter.keys()),
                    "Assigned Periods": list(load_counter.values())
                }).sort_values(by="Assigned Periods", ascending=False)

                st.bar_chart(load_df.set_index("Teacher"))
            else:
                st.info("No assignments to visualize.")

        # === MANUAL EDITING INTERFACE TRIGGER ===
        if "generated_arrangement" in st.session_state and "suggestions_df" in st.session_state:
            st.markdown("---")
            st.markdown("### 🧑‍🏫 Manual Edit Timetable")

            if "edit_queue" not in st.session_state:
                st.session_state.edit_queue = []

            if st.button("➕ Add Teacher for Editing"):
                st.session_state.edit_queue.append({
                    "teacher": "",
                    "periods": [],
                    "edits": {}
                })

            original_df = st.session_state["generated_arrangement"]
            suggestions_df = st.session_state["suggestions_df"]
            editable_df = original_df.copy()
            period_columns = [col for col in editable_df.columns if col.startswith("Period")]

            for idx, entry in enumerate(st.session_state.edit_queue):
                st.markdown(f"---\n#### 📝 Edit Entry #{idx + 1}")
                col1, col2, col3 = st.columns([4, 4, 1])

                with col3:
                    if st.button(f"🗑️", key=f"delete_entry_{idx}"):
                        st.session_state.edit_queue.pop(idx)
                        st.rerun()

                with col1:
                    absent_list = editable_df["Absent Teacher"].unique().tolist()
                    selected_teacher = st.selectbox(
                        f"👤 Absent Teacher (Entry #{idx + 1})",
                        [""] + absent_list,
                        index=absent_list.index(entry["teacher"]) + 1 if entry["teacher"] in absent_list else 0,
                        key=f"teacher_{idx}"
                    )
                    entry["teacher"] = selected_teacher

                if selected_teacher:
                    with col2:
                        teacher_df = editable_df[editable_df["Absent Teacher"] == selected_teacher]
                        
                        if teacher_df.empty:
                            st.warning(f"No timetable rows found for {selected_teacher}.")
                            continue
                        
                        teacher_periods = [int(col.split(" ")[1]) for col in period_columns if not teacher_df[col].isnull().all()]
                        selected_periods = st.multiselect(
                            f"🕘 Periods for {selected_teacher}",
                            options=sorted(teacher_periods),
                            default=entry["periods"],
                            key=f"periods_{idx}"
                        )
                        entry["periods"] = selected_periods

                    for period_num in selected_periods:
                        period_col = f"Period {period_num}"
                        current_value = teacher_df.iloc[0][period_col]
                        current_teacher = current_value.split(" (")[0].strip() if current_value else ""

                        if suggestions_df is not None and not suggestions_df.empty:
                            suggestion = suggestions_df[
                                (suggestions_df["Absent Teacher"] == selected_teacher) &
                                (suggestions_df["Period"] == period_num)
                            ]
                        else:
                            suggestion = pd.DataFrame(columns=["Absent Teacher", "Period", "Class", "Suggested Teachers"])
                        class_val = suggestion["Class"].values[0] if not suggestion.empty else "N/A"
                        suggested_teachers = suggestion["Suggested Teachers"].values[0].split(", ") if not suggestion.empty else []
                        options = [""] + suggested_teachers

                        substitute = st.selectbox(
                            f"➡️ Substitute for Period {period_num} (Class: {class_val})",
                            options=options,
                            index=options.index(current_teacher) if current_teacher in options else 0,
                            key=f"{selected_teacher}_{period_num}_{idx}"
                        )
                        entry["edits"][period_num] = f"{substitute} ({class_val})" if substitute else ""

            if st.session_state.edit_queue and st.button("🧾 Review Changes"):
                    for entry in st.session_state.edit_queue:
                        teacher = entry["teacher"]
                        if not teacher: continue
                        for period in entry["periods"]:
                            period_col = f"Period {period}"
                            new_val = entry["edits"].get(period, "")
                            editable_df.loc[editable_df["Absent Teacher"] == teacher, period_col] = new_val
                    
                    st.session_state["final_arrangement"] = editable_df
                    st.session_state["generated_arrangement"] = editable_df
                    # st.success("📋 Reviewing Changes.")
                    save_state_to_sheet(
                        date_str=today,
                        day_mode=day_mode,
                        absent_teachers=list(absent_dict.keys()),
                        reasons_dict=absence_reason_dict,
                        timetable_df=editable_df,
                        worksheet=PersistentStateWorksheet,
                        custom_periods=st.session_state.get("__meta__custom_periods", []),
                        suggestions_df=st.session_state.get("suggestions_df", pd.DataFrame(columns=["Absent Teacher", "Period", "Class", "Suggested Teachers"]))
                    )

                    st.markdown("### 🗂️ Updated Arrangement Timetable")
                    pivot_df = editable_df.drop(columns=["Absent Teacher"])
                    pivot_df.insert(0, "Class", editable_df["Absent Teacher"])
                    pivot_df.rename(columns={"Class": "Absent Teacher"}, inplace=True)
                    st.dataframe(pivot_df, use_container_width=True)

                    st.markdown("### ⚠️ Conflict Report")
                    assigned_periods = {}
                    conflict_rows = []
                    for _, row in editable_df.iterrows():
                        teacher = row["Absent Teacher"]
                        for idx, col in enumerate(period_columns, start=1):
                            assigned = row[col]
                            if pd.notna(assigned) and assigned.strip():
                                sub_teacher = assigned.split(" (")[0].strip()
                                key = (sub_teacher, idx)
                                if key in assigned_periods:
                                    conflict_rows.append({
                                        "Conflict Period": f"Period {idx}",
                                        "Teacher": sub_teacher,
                                        "Conflicting With": assigned_periods[key],
                                        "Also Assigned To": teacher
                                    })
                                else:
                                    assigned_periods[key] = teacher

                    if conflict_rows:
                        conflict_df = pd.DataFrame(conflict_rows)
                        st.error("🚨 Time-slot conflicts detected!")
                        st.dataframe(conflict_df, use_container_width=True)
                    else:
                        st.success("✅ No time-slot conflicts detected.")

                    st.markdown("### 📊 Arrangement Load per Substitute Teacher")
                    load_counter = {}
                    for _, row in editable_df.iterrows():
                        for col in period_columns:
                            assigned = row[col]
                            if pd.notna(assigned) and assigned.strip():
                                sub_teacher = assigned.split(" (")[0].strip()
                                load_counter[sub_teacher] = load_counter.get(sub_teacher, 0) + 1

                    if load_counter:
                        load_df = pd.DataFrame({
                            "Teacher": list(load_counter.keys()),
                            "Assigned Periods": list(load_counter.values())
                        }).sort_values(by="Assigned Periods", ascending=False)
                        st.bar_chart(load_df.set_index("Teacher"))
                    else:
                        st.info("No assignments to visualize.")

        # === APPLY CHANGES TO GOOGLE SHEET ===
        if "final_arrangement" in st.session_state and st.button("📤 Commit Timetable Changes"):
                final_df = st.session_state["final_arrangement"]

                today_str = datetime.today().strftime("%A, %d %B %Y")
                day_str = datetime.today().strftime("%A")

                updated = False
                for i, entry in enumerate(st.session_state.weekly_arrangements):
                    if entry["date"] == today_str:
                        st.session_state.weekly_arrangements[i] = {
                            "date": today_str,
                            "day": day_str,
                            "arrangement": final_df
                        }
                        updated = True
                        break

                if not updated:
                    st.session_state.weekly_arrangements.append({
                        "date": today_str,
                        "day": day_str,
                        "arrangement": final_df
                    })

                weekly_log_df = pd.concat([
                    log["arrangement"].assign(Date=log["date"], Day=log["day"])
                    for log in st.session_state.weekly_arrangements
                    if is_same_week(log["date"])
                ])

                try:
                    final_df = final_df.fillna("")  # Replace NaN with empty string
                    for col in final_df.select_dtypes(include=["float", "int"]).columns:
                        final_df[col] = final_df[col].replace([float("inf"), float("-inf")], 0)

                    persist_weekly_log(final_df, SPREADSHEET_ID)
                    append_to_monthly_log(final_df, SPREADSHEET_ID)

                    save_state_to_sheet(
                        date_str=today_str,
                        day_mode=day_str,
                        absent_teachers=st.session_state.get("absent_teachers", []),
                        reasons_dict=st.session_state.get("reasons_dict", {}),
                        timetable_df=final_df,
                        worksheet=PersistentStateWorksheet,
                        custom_periods=st.session_state.get("__meta__custom_periods", []),
                        suggestions_df=st.session_state.get(
                            "suggestions_df", 
                            pd.DataFrame(columns=["Absent Teacher", "Period", "Class", "Suggested Teachers"])
                        )
                    )

                    st.success("✅ Timetable successfully commited.")
                except Exception as e:
                    st.error(f"❌ Failed to update Google Sheet: {e}")

                # Prepare Excel for download
                output = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=final_df.shape[1])
                ws.cell(row=1, column=1).value = f"Arrangement for {today_str}"
                ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
                ws.cell(row=1, column=1).font = Font(bold=True, size=14)

                # Write headers
                for c_idx, col_name in enumerate(final_df.columns, start=1):
                    ws.cell(row=2, column=c_idx).value = col_name
                    ws.cell(row=2, column=c_idx).font = Font(bold=True)

                # Write data
                for r_idx, row in enumerate(final_df.itertuples(index=False), start=3):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx).value = value

                wb.save(output)
                output.seek(0)

                st.download_button(
                    label="📥 Download Excel",
                    data=output.getvalue(),
                    file_name=f"arrangement_{today_str.replace(',', '').replace(' ', '_')}.xlsx"
                )
                st.markdown("---")

# Arrangement Tracker Page
elif page == "📊 Arrangement Tracker":
    st.markdown("### 🗂️ Arrangement Tracker")
    view_option = st.radio("Select View", ["Current Week", "Last Week", "Month Wise"], horizontal=True)

    if view_option == "Current Week":
        weekly_log_df = load_weekly_log(SPREADSHEET_ID)
        if weekly_log_df.empty:
            st.info("No arrangements generated this week.")
        else:
            week_dates = get_current_week_dates()
            for date in week_dates:
                day_group = weekly_log_df[weekly_log_df["Date"] == date]
                if not day_group.empty:
                    day_name = day_group["Day"].iloc[0]
                    st.markdown(f"### 📌 {date}")
                    display_df = day_group.drop(columns=["Date", "Day"]).reset_index(drop=True)
                    st.dataframe(display_df, use_container_width=True)
                    st.markdown("---")

    elif view_option == "Last Week":
        weekly_log_df = load_weekly_log(SPREADSHEET_ID)
        if weekly_log_df.empty:
            st.info("No arrangements found for last week.")
        else:
            last_week_dates = get_last_week_dates()
            for date in last_week_dates:
                day_group = weekly_log_df[weekly_log_df["Date"].str.strip() == date]
                if not day_group.empty:
                    day_name = day_group["Day"].iloc[0]
                    st.markdown(f"### 📌 {date}")
                    display_df = day_group.drop(columns=["Date", "Day"]).reset_index(drop=True)
                    st.dataframe(display_df, use_container_width=True)
                    st.markdown("---")

    elif view_option == "Month Wise":
        today = datetime.today()
        month_name = today.strftime("%B")
        ws = get_or_create_worksheet(sheet_id=SPREADSHEET_ID, worksheet_name=f"{month_name}Log")
        month_df = load_df_from_gsheet(ws)

        if month_df.empty:
            st.info("No arrangements found for this month.")
        else:
            for date, group in month_df.groupby("Date"):
                day_name = group["Day"].iloc[0] if "Day" in group.columns else ""
                st.markdown(f"### 📌 {date}")
                display_df = group.drop(columns=["Date", "Day"], errors="ignore").reset_index(drop=True)
                st.dataframe(display_df, use_container_width=True)
                st.markdown("---")